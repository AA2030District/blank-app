import streamlit as st
import PyPDF2
import os
from openpyxl import load_workbook
from datetime import datetime
import io

# Set page title and description
st.title("DTE Solar PDF Uploader!")
st.write("Upload your DTE PDFs (all at once for a given meter) and press upload! It will give you back a spreadsheet in the ESPM format.")

def convert_to_float(value_str):
    """Convert string to float by removing commas and handling other characters."""
    try:
        cleaned_str = value_str.replace(',', '').strip()
        return float(cleaned_str)
    except (ValueError, AttributeError) as e:
        st.error(f"Error converting value '{value_str}' to float: {str(e)}")
        return 0.0

def extract_text_from_pdf(pdf_file):
    """Extract text from a PDF file."""
    try:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        st.error(f"Error processing PDF: {str(e)}")
        return None

def check_for_duplicate(ws, startdate, enddate):
    """Check if an entry with the same start and end date already exists."""
    for row in range(2, ws.max_row + 1):
        existing_start = ws.cell(row=row, column=1).value
        existing_end = ws.cell(row=row, column=2).value
        if existing_start == startdate and existing_end == enddate:
            return True
    return False

def process_pdfs(uploaded_files):
    """Process uploaded PDFs and return the output workbook."""
    try:
        # Load the template workbook
        current_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(current_dir, 'MeterConsumptionDataSpreadsheet_onsite_en (1).xlsx')
        
        if not os.path.exists(template_path):
            st.error(f"Template file not found at: {template_path}")
            return None, [], []
            
        wb = load_workbook(template_path)
        ws = wb.active
        
        processed_files = []
        skipped_files = []
        
        for uploaded_file in uploaded_files:
            try:
                # Extract text from PDF
                text = extract_text_from_pdf(uploaded_file)
                
                if text is None:
                    skipped_files.append(f"{uploaded_file.name} (failed to read PDF)")
                    continue
                
                if "Detail Charges" in text:
                    totalgen = None
                    startdate = None
                    enddate = None
                    outflow = None
                    
                    for line in text.splitlines():
                        if "GenW-W" in line:
                            totalgen = line.split()[9]
                        if "R18-kWH Outflow" in line:
                            outflow = line.split()[2].replace("KWH","")
                        if "Billing Period:" in line:
                            startdate = line.split()[4]
                            enddate = line.split()[6]
                    
                    if all([startdate, enddate, totalgen, outflow]):
                        if not check_for_duplicate(ws, startdate, enddate):
                            next_row = ws.max_row + 1
                            totalgen_float = convert_to_float(totalgen)
                            outflow_float = convert_to_float(outflow)
                            
                            ws.cell(row=next_row, column=1, value=startdate)
                            ws.cell(row=next_row, column=2, value=enddate)
                            ws.cell(row=next_row, column=3, value=totalgen_float - outflow_float)
                            ws.cell(row=next_row, column=4, value=outflow_float)
                            processed_files.append(uploaded_file.name)
                        else:
                            skipped_files.append(f"{uploaded_file.name} (duplicate)")
                            
                elif "Detail of Current Charges" in text:
                    totalgen = None
                    startdate = None
                    enddate = None
                    outflow = None
                    
                    for line in text.splitlines():
                        if "Gen Solar" in line:
                            totalgen = line.split()[2]
                        if "Service Period" in line:
                            dates = line.split()
                            startdate = dates[2] + " " + dates[3].replace(",", " ") + dates[4]
                            startdate = datetime.strptime(startdate.strip(), "%b %d %Y")
                            startdate = startdate.strftime("%m/%d/%Y")
                            enddate = dates[6] + " " + dates[7].replace(",", " ") + dates[8]
                            enddate = datetime.strptime(enddate.strip(), "%b %d %Y")
                            enddate = enddate.strftime("%m/%d/%Y")
                        if "KWH Outflow" in line:
                            outflow = line.split()[2]
                    
                    if all([startdate, enddate, totalgen, outflow]):
                        if not check_for_duplicate(ws, startdate, enddate):
                            next_row = ws.max_row + 1
                            totalgen_float = convert_to_float(totalgen)
                            outflow_float = convert_to_float(outflow)
                            
                            ws.cell(row=next_row, column=1, value=startdate)
                            ws.cell(row=next_row, column=2, value=enddate)
                            ws.cell(row=next_row, column=3, value=totalgen_float - outflow_float)
                            ws.cell(row=next_row, column=4, value=outflow_float)
                            processed_files.append(uploaded_file.name)
                        else:
                            skipped_files.append(f"{uploaded_file.name} (duplicate)")
                else:
                    skipped_files.append(f"{uploaded_file.name} (unsupported format)")
                    
            except Exception as e:
                st.error(f"Error processing {uploaded_file.name}: {str(e)}")
                skipped_files.append(f"{uploaded_file.name} (error)")
        
        return wb, processed_files, skipped_files
    except Exception as e:
        st.error(f"Error loading template: {str(e)}")
        return None, [], []

def main():
    # Only show the PDF uploader
    uploaded_files = st.file_uploader("Choose PDF files", type=['pdf'], accept_multiple_files=True)
    
    if uploaded_files:
        if st.button("Process Files"):
            with st.spinner("Processing PDF files..."):
                wb, processed_files, skipped_files = process_pdfs(uploaded_files)
                
                if wb is None:
                    st.error("Failed to process files due to template loading error.")
                    return
                
                # Save the workbook to a bytes buffer
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                # Show results
                if processed_files:
                    st.success(f"Successfully processed {len(processed_files)} files")
                    st.write("Processed files:")
                    for file in processed_files:
                        st.write(f"✓ {file}")
                
                if skipped_files:
                    st.warning("Skipped files:")
                    for file in skipped_files:
                        st.write(f"⚠ {file}")
                
                # Download button for the output file
                st.download_button(
                    label="Download Excel File",
                    data=output.getvalue(),
                    file_name="Output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

if __name__ == "__main__":
    main() 